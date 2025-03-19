"""
Application to read a file containg work shifts and produce daily schedules
    Input: Excel file starting at third row. Columns (1 based):
        1: Employee number
        3: Employee last name
        4: Employee first name
        6: Shift date
        7: Shift time "start - stop" (HH:MM - HH:MM)
    Output: Excel file(s) per week with a sheet for each weekday
        Will also have phone numbers filled in if matching reserverd (see Settings/Settings.xlsx)
        Will also append a footer to each sheet (from Settings/Footer.xlsx)
"""
from datetime import datetime, timedelta
import sys
from dataclasses import dataclass
import tkinter as tk
from tkinter import filedialog
from os.path import exists
from copy import copy
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from attr import field
import yaml
from glob import glob
from pathlib import Path
from argparse import ArgumentParser
import csv

SETTINGS_PATH_FILE = 'app.config'
APP_CONFIG_FILE = 'config.yaml'
EMPLOYEE_FILE = 'employee.csv'

def parse_path(path: str) -> str:
    return path.replace('<HOME_FOLDER>', str(Path.home())).strip()

class Config:
    _settings = None
    _employee_data = {}

    @staticmethod
    def settings() -> dict:
        if not Config._settings:
            Config._build()
        return Config._settings

    @staticmethod
    def employee_data() -> dict:
        return Config._employee_data

    @classmethod
    def _build(cls):
        with open(SETTINGS_PATH_FILE) as f:
            settings_path = f.readline()
            Config._settings = yaml.safe_load(open(f'{parse_path(settings_path)}/{APP_CONFIG_FILE}'))
        with open(parse_path(Config._settings['paths']['employee_data']) + f'/{EMPLOYEE_FILE}') as f:
            spamreader = csv.reader(f)
            Config._employee_data = {}
            for row in spamreader:
                if row[0] == 'id':
                    continue
                Config._employee_data[int(row[0])] = {
                    'phone': row[1],
                    'role': row[2],
                }


@dataclass
class WorkShift:
    """Defines a work shift including employee details and shift time"""
    employee_id: str
    date: datetime
    employee_fname: str = field(repr=False)
    employee_lname: str = field(repr=False)
    shift_time: str = field(repr=False)
    start_time: str = field(init=False)
    end_time: str = field(init=False)
    total: datetime = field(init=False, repr=False)
    employee_name: str = field(init=False)
    phone_number: int = None
    employee_role: str = ''
    dates = []

    def __post_init__(self) -> None:
        self.employee_name = f"{self.employee_fname} {self.employee_lname}"
        if self.date not in WorkShift.dates:
            WorkShift.dates.append(self.date)
            WorkShift.dates.sort()
        tmp_start, tmp_end = self.shift_time.split(' - ')
        self.start_time = datetime.strptime(f"{tmp_start}",'%H:%M')
        self.end_time = datetime.strptime(f"{tmp_end}",'%H:%M')
        # If entitled to lunch set it after defined hours
        if self.end_time - self.start_time > timedelta(hours=Config.settings()['settings']['hours_for_lunch']):
            self.lunch = self.start_time + timedelta(hours=Config.settings()['settings']['lunch_after'])
        else:
            self.lunch = None
        self.total = self.end_time - self.start_time
        if self.lunch is not None:
            self.total -= timedelta(hours=1)
        if self.employee_id in Config.employee_data().keys():
            self.phone_number = Config.employee_data()[self.employee_id].get('phone', None)
            self.employee_role = Config.employee_data()[self.employee_id].get('role', None)

    def get_list(self, time_range) -> list:
        """Returns a list of strings to describe how the linked employee works during the shift"""
        result = [self.shift_time, self.employee_name, self.phone_number]
        for time in time_range:
            tm = datetime.strptime(time, '%H:%M')
            if self.lunch is not None and\
                (tm <= self.lunch and tm > (self.lunch - timedelta(hours=1))):
                result.append('Lunch')
            elif tm >= self.start_time and tm < self.end_time:
                result.append(self.employee_role)
            else:
                result.append('FREE')
        result.append(self.total)
        return result


def copy_footer(source, destination):
    """
    Load a footer (Excel) file from source and copy content at
    end of destination (Excel) file including formatting
    """
    wb_source = load_workbook(source)
    ws_source = wb_source.active
    wb_destination = load_workbook(destination, read_only=False)

    for sheet in wb_destination.sheetnames:
        ws_destination = wb_destination[sheet]

        for row in ws_source.iter_rows():
            dest_row = ws_destination.max_row+1
            for cell in row:
                new_cell = ws_destination.cell(row=dest_row, column=cell.column)
                new_cell.value = cell.value
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)
    wb_destination.save(destination)


def read_shifts(input_files: list[str]) -> list[WorkShift]:
    all_shifts = []
    for file in input_files:
        print(f"Getting shift data from input: {file}")
        wb_in = load_workbook(file)
        ws_in = wb_in.active
        for row in ws_in.iter_rows(min_row=Config.settings()['input_format']['row_start'], values_only=True):
            if row[0] is not None:
                shift = WorkShift(
                    employee_id=row[Config.settings()['input_format']['col_id']],
                    employee_lname=row[Config.settings()['input_format']['col_last_name']],
                    employee_fname=row[Config.settings()['input_format']['col_first_name']],
                    date=row[Config.settings()['input_format']['col_date']],
                    shift_time=row[Config.settings()['input_format']['col_shift']]
                    )
                if shift is not None:
                    all_shifts.append(shift)
                else:
                    print(f"Failure to create WorkShift with input data {row}")
    print(f"Processed {len(all_shifts)} lines of shift data")
    return all_shifts


def create_schedules(shift_data: list[WorkShift], output_folder: str, footer_folder: str = None):
    print("Creating list of weeks/dates")
    week_dates = {}
    for date in WorkShift.dates:
        week_no = datetime.strptime(date,'%Y-%m-%d').isocalendar()[1]
        if week_no not in week_dates:
            week_dates[week_no] = []
        week_dates[week_no].append(date)

    for week in week_dates.items():
        print(f"Creating new output workbook for week {week[0]}")
        wb_out = Workbook()
        wb_out.remove(wb_out['Sheet'])
        out_file_name = f"{output_folder}/Vecka {week[0]}.xlsx"

        for date in week_dates[week[0]]:
            print(f"Creating temporary shift data for date: {date}")
            current_day_shifts: list[WorkShift] = []
            current_day_earliest = None # Current day's earliest start
            current_day_latest = None # Current day's latest end
            for shift in shift_data:
                if shift.date == date:
                    current_day_shifts.append(shift)
                    if current_day_earliest is None or current_day_earliest > shift.start_time:
                        current_day_earliest = shift.start_time
                    if current_day_latest is None or current_day_latest < shift.end_time:
                        current_day_latest = shift.end_time
            print(f"- Sorting {len(current_day_shifts)} shifts for current day")
            current_day_shifts.sort(key=lambda x: x.start_time)

            print(f"- Generating time ranges for current date ({current_day_earliest.strftime('%H:%M')}-{current_day_latest.strftime('%H:%M')})")
            time_range = [] # Start+End of each hour (for simple print in heading)
            time_range_simple = [] # Start of each hour (for simple compare with datetime objects)
            shift_start = current_day_earliest
            while shift_start < current_day_latest:
                shift_end = shift_start + timedelta(hours=1)
                if shift_end > current_day_latest:
                    shift_end = current_day_latest
                time_range.append(shift_start.strftime('%H:%M') + '-' + shift_end.strftime('%H:%M'))
                time_range_simple.append(shift_start.strftime('%H:%M'))
                shift_start = shift_end

            print("- Excel: Creating new sheet for current date")
            ws_out = wb_out.create_sheet(date)
            ws_out.title = datetime.strptime(date,'%Y-%m-%d').strftime('%A')

            print("- Excel: Printing day and date header")
            ws_out.append([datetime.strptime(date,'%Y-%m-%d').strftime('%A') + f" - {date}"])

            print("- Excel: Printing headings for current date")
            ws_out.append(['Arbetstid','Namn','Tele'] + time_range)

            print("- Excel: Adding shifts to current date")
            for shift in current_day_shifts:
                ws_out.append(shift.get_list(time_range_simple))

            print("- Excel: Printing headings at bottom as well + total time formula")
            ws_out.append(['Arbetstid','Namn','Tele'] + time_range)

            print("- Excel: Deleting shift columns before 10:00")
            for tm in time_range_simple:
                if datetime.strptime(tm, '%H:%M') < datetime.strptime('10:00','%H:%M'):
                    ws_out.delete_cols(4) # The 4th column is the first containing shift times
                else:
                    break

            print("- Excel: Printing total time formula")
            sum_total_time_formula = f"=SUM({get_column_letter(ws_out.max_column) + str(ws_out.min_row)}:{get_column_letter(ws_out.max_column) + str(ws_out.max_row - 1)})"
            sum_cell = ws_out.cell(row=ws_out.max_row, column=ws_out.max_column)
            sum_cell.value = sum_total_time_formula
            sum_cell.number_format = '[h]:mm;@'

            print("- Excel: Formatting cells")
            thin = Side(border_style="thin", color="000000")
            for row in ws_out.iter_rows():
                for cell in row:
                    # Titel row (day + date)
                    if cell.row == 1:
                        ws_out.merge_cells(start_row=1, end_row=1, start_column=1,\
                            end_column=ws_out.max_column)
                        ws_out.row_dimensions[cell.row].height = 25
                        cell.alignment = Alignment(horizontal="center")
                        cell.fill = PatternFill("solid", fgColor="A3A3A3")
                        cell.font = Font(color="FFFFFF", size=20)
                    # Borders of whole table
                    if cell.row > 1:
                        cell.border = Border(top=thin,left=thin,bottom=thin,right=thin)
                    # Bolden columns/rows
                    if (cell.column <= 3 and cell.row > 1) or\
                        cell.row == 2 or cell.row == ws_out.max_row:
                        cell.font = Font(bold=True)
                    # Column width, text alignment and fill (static)
                    if cell.column == 1:
                        ws_out.column_dimensions[get_column_letter(cell.column)].width = 15
                    if cell.column == 2:
                        ws_out.column_dimensions[get_column_letter(cell.column)].width = 25
                        if cell.row > 2 and cell.row < ws_out.max_row:
                            cell.fill = PatternFill("solid", fgColor="FFE6B3")
                    if cell.column == 3:
                        cell.alignment = Alignment(horizontal="center")
                        ws_out.column_dimensions[get_column_letter(cell.column)].width = 10
                    if cell.column > 3:
                        cell.alignment = Alignment(horizontal="center")
                        ws_out.column_dimensions[get_column_letter(cell.column)].width = 15
                    if cell.row == 2 or cell.row == ws_out.max_row:
                        cell.fill = PatternFill("solid", fgColor="99CCFF")
                    # Cell formatting/fill depending on value
                    if cell.value == 'FREE':
                        cell.value = ''
                        cell.fill = PatternFill("solid", fgColor="C0C0C0")
                    elif cell.value == 'Lunch':
                        cell.fill = PatternFill("solid", fgColor="FFFFCC")

        print(f"Writing NEW FILE to: {out_file_name}")
        wb_out.save(out_file_name)

        if footer_folder and exists(footer_folder):
            for footer_file in glob(f'{footer_folder}/*.xlsx'):
                print(f"Found footer file: {footer_file}, adding to all sheets in {out_file_name}")
                copy_footer(footer_file, out_file_name)


def main():
    """Main application"""
    parser = ArgumentParser(prog='Schedule helper')
    parser.add_argument('--input-file', '-i', help='Input xlsx file with shift times', required=False)
    args = parser.parse_args()

    if not args.input_file:
        print("Getting input file path from user")
        root = tk.Tk()
        root.withdraw()
        args.input_file = filedialog.askopenfilenames(
            title="Pick input file(s)",
            filetypes=[('Excel files','.xlsx')],
            initialdir=parse_path(Config.settings()['paths'].get('default', '.'))
        )
        if len(args.input_file) == 0:
            print("Cancelled by user")
            sys.exit()

    all_shifts = read_shifts(input_files=args.input_file)
    create_schedules(
        shift_data=all_shifts,
        output_folder=parse_path(Config.settings()['paths'].get('output', '.')),
        footer_folder=parse_path(Config.settings()['paths'].get('footers', None))
    )


if __name__ == "__main__":
    main()
