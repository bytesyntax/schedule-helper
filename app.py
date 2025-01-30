"""
Application to read a file containg work shifts and produce daily schedules
    Input: Excel file starting at third row. Columns (1 based):
        1: Employee number
        3: Employee last name
        4: Employee first name
        6: Shift date
        7: Shift time "start - stop" (HH:MM - HH:MM)
    Output: Excel file(s) per week with a sheet for each weekday
        Will also have phone numbers filled in if matching reserverd (see Settings/Settings.xlsx)
        Will also append a footer to each sheet (from Settings/Footer.xlsx)
"""
from datetime import datetime, timedelta
import sys
from dataclasses import dataclass
import tkinter as tk
from tkinter import filedialog
from os.path import exists
from copy import copy
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from attr import field

SETTINGS_FILE = 'Settings/Settings.xlsx'

@dataclass
class EmployeeSettings:
    """Used for reading settings such as reserved phone numbers and stations"""
    # Settings that will be parsed from settings file
    setting_type = {
        'phone': 'dict',
        'role': 'dict',
        'lunch_after': 'value',
        'hours_for_lunch': 'value',
        'col_fname': 'value',
        'col_lname': 'value',
        'col_id': 'value',
        'col_date': 'value',
        'col_shifttime': 'value',
        'row_input_start': 'value',
        'footer_file': 'value',
        'output_folder': 'value',
        'open_default_folder': 'value',
    }
    settings = {}
    if exists(SETTINGS_FILE):
        wb = load_workbook(SETTINGS_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            print(row)
            st = row[0].strip()
            if st not in setting_type:
                print(f'Unknown setting {st}')
                continue
            if setting_type[st] == 'dict':
                if st not in settings:
                    settings[st] = {}
                settings[st][row[1]] = row[2]
            elif setting_type[st] == 'value':
                settings[st] = row[2]
    else:
        raise SystemExit(f'The settings file "{SETTINGS_FILE}" does not exist or is not readable')

    print(settings)

    @classmethod
    def get_number(cls,employee_id) -> int:
        """Retrieve reserverd phone number for employee id"""
        if employee_id in cls.settings['phone']:
            return cls.settings['phone'][employee_id]
        else:
            return None

    @classmethod
    def get_role(cls, employee_id) -> str:
        """Retrieve reserverd station/role for employee id"""
        if employee_id in cls.settings['role']:
            return cls.settings['role'][employee_id]
        else:
            return None


@dataclass
class WorkShift:
    """Defines a work shift including employee details and shift time"""
    employee_id: str
    date: datetime
    employee_fname: str = field(repr=False)
    employee_lname: str = field(repr=False)
    shift_time: str = field(repr=False)

    start_time: str = field(init=False)
    end_time: str = field(init=False)
    total: datetime = field(init=False, repr=False)
    employee_name: str = field(init=False)
    phone_number: int = field(init=False)
    employee_role: str = field(init=False)

    dates = []

    def __post_init__(self) -> None:
        self.employee_name = f"{self.employee_fname} {self.employee_lname}"
        if self.date not in WorkShift.dates:
            WorkShift.dates.append(self.date)
            WorkShift.dates.sort()
        tmp_start, tmp_end = self.shift_time.split(' - ')
        self.start_time = datetime.strptime(f"{tmp_start}",'%H:%M')
        self.end_time = datetime.strptime(f"{tmp_end}",'%H:%M')
        # If entitled to lunch set it after defined hours
        if self.end_time - self.start_time > timedelta(hours=EmployeeSettings.settings['hours_for_lunch']):
            self.lunch = self.start_time + timedelta(hours=EmployeeSettings.settings['lunch_after'])
        else:
            self.lunch = None
        self.total = self.end_time - self.start_time
        if self.lunch is not None:
            self.total -= timedelta(hours=1)
        self.phone_number = EmployeeSettings.get_number(self.employee_id)
        self.employee_role = EmployeeSettings.get_role(self.employee_id)

    def map_work_time(self, time_range) -> list:
        """Returns a list of strings to describe how the linked emplyee works during the shift"""
        result = []
        for time in time_range:
            tm = datetime.strptime(time, '%H:%M')
            if self.lunch is not None and\
                (tm <= self.lunch and tm > (self.lunch - timedelta(hours=1))):
                result.append('Lunch')
            elif tm >= self.start_time and tm < self.end_time:
                result.append(self.employee_role)
            else:
                result.append('FREE')
        return result

def copy_footer(source, destination):
    """
    Load a footer (Excel) file from source and copy content at
    end of destination (Excel) file including formatting
    """
    wb_source = load_workbook(source)
    ws_source = wb_source.active
    wb_destination = load_workbook(destination, read_only=False)

    for sheet in wb_destination.sheetnames:
        ws_destination = wb_destination[sheet]

        for row in ws_source.iter_rows():
            dest_row = ws_destination.max_row+1
            for cell in row:
                new_cell = ws_destination.cell(row=dest_row, column=cell.column)
                new_cell.value = cell.value
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)

    wb_destination.save(destination)


def main():
    """Main application"""
    print("Getting input file path from user")
    root = tk.Tk()
    root.withdraw()

    data_in = filedialog.askopenfilenames(
        title="Pick input file(s)",
        filetypes=[('Excel files','.xlsx')],
        initialdir=EmployeeSettings.settings['open_default_folder']
        )
    if len(data_in) == 0:
        print("Cancelled by user")
        sys.exit()

    all_shifts = []
    for file in data_in:
        print(f"Getting shift data from input: {file}")
        wb_in = load_workbook(file)
        ws_in = wb_in.active
        for row in ws_in.iter_rows(min_row=EmployeeSettings.settings['row_input_start'], values_only=True):
            if row[0] is not None:
                shift = WorkShift(
                    employee_id=row[EmployeeSettings.settings['col_id']],
                    employee_lname=row[EmployeeSettings.settings['col_lname']],
                    employee_fname=row[EmployeeSettings.settings['col_fname']],
                    date=row[EmployeeSettings.settings['col_date']],
                    shift_time=row[EmployeeSettings.settings['col_shifttime']]
                    )
                if shift is not None:
                    all_shifts.append(shift)
                else:
                    print(f"Failure to create WorkShift with input data {row}")

    print(f"Processed {len(all_shifts)} lines of shift data")

    print("Creating list of weeks/dates")
    week_dates = {}
    for date in WorkShift.dates:
        week_no = datetime.strptime(date,'%Y-%m-%d').isocalendar()[1]
        if week_no not in week_dates:
            week_dates[week_no] = []
        week_dates[week_no].append(date)

    for week in week_dates.items():
        print(f"Creating new output workbook for week {week[0]}")
        wb_out = Workbook()
        wb_out.remove(wb_out['Sheet'])
        out_file_name = f"{EmployeeSettings.settings['output_folder']}Vecka {week[0]}.xlsx"

        for date in week_dates[week[0]]:
            print(f"Creating temporary shift data for date: {date}")
            current_day_shifts = []
            current_day_earliest = None # Current day's earliest start
            current_day_latest = None # Current day's latest end
            for shift in all_shifts:
                if shift.date == date:
                    current_day_shifts.append(shift)
                    if current_day_earliest is None or current_day_earliest > shift.start_time:
                        current_day_earliest = shift.start_time
                    if current_day_latest is None or current_day_latest < shift.end_time:
                        current_day_latest = shift.end_time
            print(f"- Sorting {len(current_day_shifts)} shifts for current day")
            current_day_shifts.sort(key=lambda x: x.start_time)

            print(f"- Generating time ranges for current date (\
                {current_day_earliest.strftime('%H:%M')}-{current_day_latest.strftime('%H:%M')})")
            time_range = [] # Start+End of each hour (for simple print in heading)
            time_range_simple = [] # Start of each hour (for simple compare with datetime objects)
            shift_start = current_day_earliest
            while shift_start < current_day_latest:
                shift_end = shift_start + timedelta(hours=1)
                if shift_end > current_day_latest:
                    shift_end = current_day_latest
                time_range.append(shift_start.strftime('%H:%M') + '-' + shift_end.strftime('%H:%M'))
                time_range_simple.append(shift_start.strftime('%H:%M'))
                shift_start = shift_end

            print("- Excel: Creating new sheet for current date")
            ws_out = wb_out.create_sheet(date)
            ws_out.title = datetime.strptime(date,'%Y-%m-%d').strftime('%A')

            print("- Excel: Printing day and date header")
            ws_out.append([datetime.strptime(date,'%Y-%m-%d').strftime('%A') + f" - {date}"])

            print("- Excel: Printing headings for current date")
            ws_out.append(['Arbetstid','Namn','Tele'] + time_range)

            print("- Excel: Adding shifts to current date")
            for shift in current_day_shifts:
                ws_out.append([shift.shift_time, shift.employee_name, shift.phone_number]\
                    + shift.map_work_time(time_range_simple) + [shift.total])

            print("- Excel: Printing headings at bottom as well + total time formula")
            ws_out.append(['Arbetstid','Namn','Tele'] + time_range)

            print("- Excel: Deleting shift columns before 10:00")
            for tm in time_range_simple:
                if datetime.strptime(tm, '%H:%M') < datetime.strptime('10:00','%H:%M'):
                    ws_out.delete_cols(4) # The 4th column is the first containing shift times
                else:
                    break

            print("- Excel: Printing total time formula")
            sum_total_time_formula = f"=SUM({get_column_letter(ws_out.max_column) + str(ws_out.min_row)}:{get_column_letter(ws_out.max_column) + str(ws_out.max_row - 1)})"
            sum_cell = ws_out.cell(row=ws_out.max_row, column=ws_out.max_column)
            sum_cell.value = sum_total_time_formula
            sum_cell.number_format = '[h]:mm;@'

            print("- Excel: Formatting cells")
            thin = Side(border_style="thin", color="000000")
            for row in ws_out.iter_rows():
                for cell in row:
                    # Titel row (day + date)
                    if cell.row == 1:
                        ws_out.merge_cells(start_row=1, end_row=1, start_column=1,\
                            end_column=ws_out.max_column)
                        ws_out.row_dimensions[cell.row].height = 25
                        cell.alignment = Alignment(horizontal="center")
                        cell.fill = PatternFill("solid", fgColor="A3A3A3")
                        cell.font = Font(color="FFFFFF", size=20)
                    # Borders of whole table
                    if cell.row > 1:
                        cell.border = Border(top=thin,left=thin,bottom=thin,right=thin)
                    # Bolden columns/rows
                    if (cell.column <= 3 and cell.row > 1) or\
                        cell.row == 2 or cell.row == ws_out.max_row:
                        cell.font = Font(bold=True)
                    # Column width, text alignment and fill (static)
                    if cell.column == 1:
                        ws_out.column_dimensions[get_column_letter(cell.column)].width = 15
                    if cell.column == 2:
                        ws_out.column_dimensions[get_column_letter(cell.column)].width = 25
                        if cell.row > 2 and cell.row < ws_out.max_row:
                            cell.fill = PatternFill("solid", fgColor="FFE6B3")
                    if cell.column == 3:
                        cell.alignment = Alignment(horizontal="center")
                        ws_out.column_dimensions[get_column_letter(cell.column)].width = 10
                    if cell.column > 3:
                        cell.alignment = Alignment(horizontal="center")
                        ws_out.column_dimensions[get_column_letter(cell.column)].width = 15
                    if cell.row == 2 or cell.row == ws_out.max_row:
                        cell.fill = PatternFill("solid", fgColor="99CCFF")
                    # Cell formatting/fill depending on value
                    if cell.value == 'FREE':
                        cell.value = ''
                        cell.fill = PatternFill("solid", fgColor="C0C0C0")
                    elif cell.value == 'Lunch':
                        cell.fill = PatternFill("solid", fgColor="FFFFCC")

        print(f"Writing NEW FILE to: {out_file_name}")
        wb_out.save(out_file_name)

        if exists(EmployeeSettings.settings['footer_file']):
            print(f"Found footer file: {EmployeeSettings.settings['footer_file']}, adding to all sheets in {out_file_name}")
            copy_footer(EmployeeSettings.settings['footer_file'], out_file_name)

if __name__ == "__main__":
    main()
